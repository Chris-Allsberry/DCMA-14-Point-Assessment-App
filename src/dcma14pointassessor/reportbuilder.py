from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    NamedStyle,
    Font,
    PatternFill,
    Border,
    Color,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter

from .validation.validation_classes import ValidationResult

class ReportBuilder:
    def __init__(self, validation_results: list[ValidationResult]):
        self.results = validation_results
        self.wb = Workbook()
        self.summary_sheet = None
        self.detail_sheet = None

    def __data_transformer(self,data):
        output = []
        output.append(list(data[0].keys()))
        for i in data:
            output.append(list(i.values()))
        return output

    def __add_summary_data(self):
        self.summary_sheet = self.wb.active
        self.summary_sheet.title = "Summary"
        all_data = []
        for res in self.results:
            all_data.append(res.to_summary_dict())
        summary_tranformed = self.__data_transformer(all_data)
        for row in summary_tranformed:
            self.summary_sheet.append(row)

    def __add_details_data(self):
        self.details_sheet = self.wb.create_sheet("Details")
        all_data = []
        for r in self.results:
            for rd in r.create_details_list():
                all_data.append(rd)
        details_transformed = self.__data_transformer(all_data)
        for row in details_transformed:
            self.details_sheet.append(row)

    def __create_named_styles(self):
            # Summary Header
            style_summary_header = NamedStyle(name="SummaryHeader")
            style_summary_header.font = Font(bold=True, size=14, color="60497A")
            style_summary_header.fill = PatternFill(patternType="solid", fgColor="B8CCE4")
            # bd = Side(style='thick', color='000000')
            # summary_header.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            self.wb.add_named_style(style_summary_header)

            # Pass
            style_pass = NamedStyle(name="Pass")
            style_pass.fill = PatternFill(patternType="solid", fgColor="00008000")
            style_pass.font = Font(bold=True, color="00FFFFFF")
            self.wb.add_named_style(style_pass)

            # Fail
            style_fail = NamedStyle(name="Fail")
            style_fail.fill = PatternFill(patternType="solid", fgColor="00FF0000")
            style_fail.font = Font(bold=True, color="00FFFF00")
            self.wb.add_named_style(style_fail)

            # N/A
            style_na = NamedStyle(name="NA")
            style_na.fill = PatternFill(patternType="solid", fgColor="00003366")
            style_na.font = Font(bold=True, color="00FFFF00")
            self.wb.add_named_style(style_na)

            # Details Header
            style_detail_header = NamedStyle(name="DetailsHeader")
            style_detail_header.font = Font(bold=True, size=14, color="000000")
            style_detail_header.fill = PatternFill(patternType="solid", fgColor="FABF8F")
            self.wb.add_named_style(style_detail_header)

            # Project Info Field
            style_project_info_field = NamedStyle(name="ProjectInfoField")
            style_project_info_field.font = Font(bold=True, size=14, color="FFFFFF")
            style_project_info_field.fill = PatternFill(patternType="solid", fgColor="963634")
            self.wb.add_named_style(style_project_info_field)

            # Project Info Value
            style_project_info_value = NamedStyle(name="ProjectInfoValue")
            style_project_info_value.font = Font(bold=True, size=11, color="000000")
            style_project_info_value.fill = PatternFill(patternType="solid", fgColor="E6B8B7")
            self.wb.add_named_style(style_project_info_value)

    def __apply_summary_styling(self):
        # Set Summary header style
        for col in self.summary_sheet["A1":"F1"]:
            for cell in col:
                cell.style = "SummaryHeader"

        # Set Summary Freeze Panes
        self.summary_sheet.freeze_panes = "A2"

        # Set the Result Column
        for cell in self.summary_sheet["A"]:
            if cell.value == "Pass":
                cell.style = "Good"
            elif cell.value == "Fail":
                cell.style = "Bad"
            elif cell.value == 'N/A':
                cell.style = "NA"

        # Set the Type Column
        for cell in self.summary_sheet["D"]:
            if cell.value == "Error":
                cell.style = "Warning Text"
            elif cell.value == "Warning":
                cell.style = "Neutral"

        # Set Alignment for Summary Sheet
        for col in self.summary_sheet["A":"D"]:
            for cell in col:
                cell.alignment = Alignment(horizontal="center")
        for col in self.summary_sheet["E":"F"]:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, horizontal="center")

        # Set Summary Tab Column Widths
        for col in self.summary_sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            if col_letter == "E":
                set_width = 75
            elif col_letter == "F":
                set_width = 50
            else:
                set_width = max_len * 1.2
            self.summary_sheet.column_dimensions[col_letter].width = set_width

    def __apply_detail_styling(self):
        # Set Details Header
        for col in self.details_sheet["A1":"G1"]:
            for cell in col:
                cell.style = "DetailsHeader"

        # Set Details Freeze Pane
        self.details_sheet.freeze_panes = "A2"
        
        # Set the Details Type Column
        for cell in self.details_sheet["C"]:
            if cell.value == "Error":
                cell.style = "Warning Text"
            elif cell.value == "Warning":
                cell.style = "Neutral"

        # Set Wrap Text for Details Sheet
        for col in self.details_sheet["A":"F"]:
            for cell in col:
                cell.alignment = Alignment(horizontal="center")
        for cell in self.details_sheet["G"]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        # Set Details Tab Column Widths
        for col in self.details_sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            if col_letter == "D":
                self.details_sheet.column_dimensions[col_letter].hidden = True
            elif col_letter == "G":
                set_width = 50
            else:
                set_width = max_len * 1.3
            self.details_sheet.column_dimensions[col_letter].width = set_width

    def create_xlsx(self) -> bytes:
        self.__add_summary_data()
        self.__add_details_data()
        self.__create_named_styles()
        self.__apply_summary_styling()
        self.__apply_detail_styling()
        buffer = BytesIO()
        self.wb.save(buffer)
        return buffer.getvalue()